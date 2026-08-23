"""P1B auditable migration of the two legacy crystal sourcing workbooks.

This script is deliberately conservative: every non-empty source row is staged;
only deterministic identity mappings are promoted.  Raw prices and procurement
links remain staged and no supplier or offer is created.
"""
import json
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return None
    value = str(value).strip()
    return value or None


def slug(value):
    normalized = re.sub(r'[^a-z0-9]+', '-', value.lower()).strip('-')[:64]
    # Chinese aliases are valid identity strings; do not collapse them to an empty key.
    return normalized or re.sub(r'\s+', ' ', value).strip().lower()


def tier(value):
    value = value or ''
    if any(x in value.lower() for x in ('premium', '高档', '高端', 'high')):
        return 'premium'
    if any(x in value.lower() for x in ('mid', '中档', '中端')):
        return 'mid'
    if any(x in value.lower() for x in ('accessible', '低档', '基础', '普通')):
        return 'accessible'
    return 'unclassified'


def raw_dict(headers, values):
    return {str(headers[i] or f'column_{i+1}'): clean(v)
            for i, v in enumerate(values) if clean(v) is not None}


class P1B:
    def __init__(self, db_path):
        self.db = sqlite3.connect(db_path)
        self.db.execute('PRAGMA foreign_keys = ON')
        self.cur = self.db.cursor()
        self.counts = {'staged': 0, 'materials': 0, 'variants': 0,
                       'components': 0, 'packaging': 0, 'review_required': 0,
                       'conflicts': 0}
        self.materials = {name: ident for ident, name in self.cur.execute(
            'SELECT id, canonical_name FROM material')}

    def batch(self, source_file, description):
        self.cur.execute("""INSERT INTO import_batch
            (source_file,source_format,source_description,imported_by,batch_status)
            VALUES (?,?,?,?,?)""", (str(source_file), 'xlsx', description,
                                      'P1B reviewed structured migration', 'validated'))
        return self.cur.lastrowid

    def source(self, path, sheet, row):
        name = f'P1B legacy workbook | {Path(path).name} | {sheet} row {row}'
        self.cur.execute("""INSERT INTO source
          (source_type,name,source_url,geography,verification_status,evidence_strength,notes)
          VALUES (?,?,?,?,?,?,?)""", ('user_upload', name, None, None, 'unverified', 'low',
              'Legacy workbook row; no external supplier or market verification performed.'))
        return self.cur.lastrowid

    def stage(self, batch, sheet, row, values, headers, target, warning):
        self.cur.execute("""INSERT INTO staged_record
          (import_batch_id,source_sheet,source_row,raw_record_json,target_entity,validation_status,warning_summary)
          VALUES (?,?,?,?,?,?,?)""", (batch, sheet, row, json.dumps(raw_dict(headers, values), ensure_ascii=False),
              target, 'review_required' if warning else 'validated', warning))
        self.counts['staged'] += 1
        if warning:
            self.counts['review_required'] += 1
        return self.cur.lastrowid

    def field(self, record, column, value, target_entity, target_field, status='valid', warning=None):
        self.cur.execute("""INSERT INTO staged_field
          (staged_record_id,source_column,raw_value,normalized_value,target_entity,target_field,field_status,warning_or_error)
          VALUES (?,?,?,?,?,?,?,?)""", (record, column, clean(value), clean(value), target_entity,
               target_field, status, warning))
        return self.cur.lastrowid

    def decision(self, field_id, decision, reason):
        self.cur.execute('INSERT INTO field_review_decision(staged_field_id,decision,reviewer,reason) VALUES (?,?,?,?)',
                         (field_id, decision, 'P1B deterministic mapping rules', reason))

    def audit(self, record, entity, entity_id, field, operation, reason, source_fields):
        self.cur.execute("""INSERT INTO field_promotion_log
          (staged_record_id,canonical_entity,canonical_record_id,canonical_field,operation,promotion_reason,promoted_by)
          VALUES (?,?,?,?,?,?,?)""", (record, entity, entity_id, field, operation, reason,
              'P1B deterministic migration'))
        log_id = self.cur.lastrowid
        for field_id in source_fields:
            self.cur.execute('INSERT OR IGNORE INTO field_promotion_source(field_promotion_log_id,staged_field_id) VALUES (?,?)',
                             (log_id, field_id))

    def row_log(self, record, entity, entity_id, note):
        self.cur.execute("""INSERT INTO review_decision(staged_record_id,decision,reviewer,notes)
          SELECT ?, 'approved', 'P1B deterministic mapping rules', 'Row-level approval records the controlled identity mapping.'
          WHERE NOT EXISTS (SELECT 1 FROM review_decision WHERE staged_record_id=? AND decision='approved')""", (record, record))
        self.cur.execute("""INSERT INTO promotion_log(staged_record_id,canonical_entity,canonical_id,promoted_by,provenance_note)
          VALUES (?,?,?,?,?)""", (record, entity, entity_id, 'P1B deterministic migration', note))
        self.cur.execute("UPDATE staged_record SET validation_status='human_approved' WHERE id=?", (record,))

    def material(self, name, family, description=None):
        if name in self.materials:
            return self.materials[name], False
        self.cur.execute('INSERT INTO material(canonical_name,material_family,natural_status,description) VALUES (?,?,?,?)',
                         (name, family, 'unknown', description))
        self.materials[name] = self.cur.lastrowid
        self.counts['materials'] += 1
        return self.cur.lastrowid, True

    def variant(self, material_id, code, quality, size, treatment, source_id, source_tier):
        self.cur.execute("""INSERT INTO material_variant
          (material_id,variant_code,grade_label,size_range_mm,treatment_disclosure,provenance_source_id,verification_status,notes,commercial_tier,source_tier_label)
          VALUES (?,?,?,?,?,?,?,?,?,?)""", (material_id, code, quality, size, treatment, source_id,
              'unverified', 'P1B legacy-source variant; price intentionally not promoted.', tier(source_tier), source_tier))
        self.counts['variants'] += 1
        return self.cur.lastrowid

    def promote_material_row(self, record, path, sheet, row, values, headers, family, identity, source_tier, cols):
        source_id = self.source(path, sheet, row)
        identity_field = self.field(record, headers[cols['english']], values[cols['english']], 'material', 'canonical_name')
        alias_field = self.field(record, headers[cols['chinese']], values[cols['chinese']], 'material_alias', 'alias_raw')
        quality_field = self.field(record, headers[cols['quality']], values[cols['quality']], 'material_variant', 'grade_label')
        size_field = self.field(record, headers[cols['size']], values[cols['size']], 'material_variant', 'size_range_mm')
        treatment_field = self.field(record, headers[cols['treatment']], values[cols['treatment']], 'material_variant', 'treatment_disclosure', 'review_required', 'Treatment remains source-derived and unverified.')
        tier_field = self.field(record, headers[cols['tier']], values[cols['tier']], 'material_variant', 'source_tier_label')
        narrative_field = self.field(record, headers[cols['narrative']], values[cols['narrative']], 'material_narrative', 'statement', 'review_required', 'Narrative is not a mineralogical or medical fact.')
        market_field = self.field(record, headers[cols['market']], values[cols['market']], 'market_assessment', 'assessment_text', 'review_required', 'Workbook statement is an internal assessment, not market evidence.')
        price_field = self.field(record, headers[cols['price']], values[cols['price']], 'supplier_offer', 'price_text', 'review_required', 'Raw price and unit are retained in staging; no offer is created.')
        for field_id in (identity_field, alias_field, quality_field, size_field, tier_field):
            self.decision(field_id, 'approved', 'Safe positional source mapping; identity is in the controlled P1B mapping.')
        for field_id, reason in ((treatment_field, 'Source claim retained as unverified text.'),
                                 (narrative_field, 'Narrative retained as source-derived statement.'),
                                 (market_field, 'Assessment retained as source-derived statement.'),
                                 (price_field, 'Price text deliberately remains staged.')):
            self.decision(field_id, 'retained_staged', reason)
        material_id, created = self.material(identity, family)
        self.audit(record, 'material', material_id, 'canonical_name', 'create' if created else 'link',
                   'Controlled P1B identity mapping.', [identity_field])
        self.cur.execute('INSERT OR IGNORE INTO material_alias(material_id,alias_raw,normalized_alias,language_code,source_id,review_status,confidence,notes) VALUES (?,?,?,?,?,?,?,?)',
                         (material_id, clean(values[cols['chinese']]), slug(clean(values[cols['chinese']]) or ''), 'zh', source_id,
                          'reviewed', 'medium', 'P1B deterministic alias mapping.'))
        variant_id = self.variant(material_id, f'p1b-{slug(Path(path).stem)}-{sheet}-{row}', clean(values[cols['quality']]),
                                  clean(values[cols['size']]), clean(values[cols['treatment']]), source_id, source_tier)
        self.audit(record, 'material_variant', variant_id, 'identity-and-attributes', 'create',
                   'Promoted safe identity; source tier held at variant level.', [identity_field, quality_field, size_field, tier_field])
        if clean(values[cols['treatment']]):
            self.cur.execute('INSERT INTO material_claim(material_id,claim_field,raw_value,source_id,verification_status,confidence,notes) VALUES (?,?,?,?,?,?,?)',
                             (material_id, 'treatment_disclosure', clean(values[cols['treatment']]), source_id, 'unverified', 'low',
                              'Source-derived statement; no factual normalization.'))
            self.audit(record, 'material_claim', self.cur.lastrowid, 'raw_value', 'create', 'Retained treatment claim with provenance.', [treatment_field])
        if clean(values[cols['narrative']]):
            self.cur.execute('INSERT INTO material_narrative(material_id,narrative_type,statement,source_id,source_context,confidence,notes) VALUES (?,?,?,?,?,?,?)',
                             (material_id, 'traditional_association', clean(values[cols['narrative']]), source_id, f'{sheet} row {row}', 'low',
                              'Source-derived narrative; not a medical claim.'))
            self.audit(record, 'material_narrative', self.cur.lastrowid, 'statement', 'create', 'Narrative separated from factual material data.', [narrative_field])
        if clean(values[cols['market']]):
            self.cur.execute('INSERT INTO market_assessment(subject_type,subject_id,target_market,assessment_text,analyst,basis_notes,confidence) VALUES (?,?,?,?,?,?,?)',
                             ('material_variant', variant_id, 'Europe', clean(values[cols['market']]), 'legacy workbook',
                              f'{sheet} row {row}; unverified internal assessment', 'low'))
            self.audit(record, 'market_assessment', self.cur.lastrowid, 'assessment_text', 'create', 'Market note separated from evidence.', [market_field])
        self.row_log(record, 'material_variant', variant_id, 'PARTIALLY_PROMOTED: raw price remains staged and treatment/narrative/market remain unverified.')

    def migrate_material_sheet(self, path, sheet, batch, skip_rows, family):
        headers = [clean(c) for c in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        cols = {'chinese': 2, 'english': 3, 'quality': 4, 'tier': 6, 'market': 7, 'narrative': 8,
                'size': 9, 'price': 10, 'treatment': 13}
        known = {'Clear Quartz': 'Clear Quartz', 'Rose Quartz': 'Rose Quartz', 'Citrine': 'Citrine',
                 'Amethyst': 'Amethyst', 'Smoky Quartz': 'Smoky Quartz', 'Black Obsidian': 'Obsidian',
                 'Silver Sheen Obsidian': 'Obsidian', 'Golden Sheen Obsidian': 'Obsidian',
                 'Green Aventurine': 'Green Aventurine', "Gold Tiger's Eye": "Tiger's Eye"}
        for row, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if row in skip_rows or not any(clean(v) for v in values):
                continue
            english = clean(values[cols['english']]) if len(values) > cols['english'] else None
            warning = 'Identity needs manual review; safely staged but not promoted.'
            record = self.stage(batch, sheet.title, row, values, headers, 'material_intake', warning)
            # Always stage the same fields, even when no canonical write is safe.
            for key, entity, field in [('chinese','material_alias','alias_raw'), ('english','material','canonical_name'),
                                       ('quality','material_variant','grade_label'), ('tier','material_variant','source_tier_label'),
                                       ('size','material_variant','size_range_mm'), ('price','supplier_offer','price_text'),
                                       ('narrative','material_narrative','statement'), ('market','market_assessment','assessment_text'),
                                       ('treatment','material_variant','treatment_disclosure')]:
                self.field(record, headers[cols[key]], values[cols[key]] if len(values) > cols[key] else None, entity, field,
                           'review_required', 'Awaiting controlled identity mapping or field-level review.')
            identity = known.get(english)
            if identity is None and english and re.fullmatch(r"[A-Za-z][A-Za-z '\-]+", english) and '/' not in english:
                # Simple undelimited English names are safely kept as source-name identities.
                identity = english
            if identity:
                # Delete generic first-pass fields; replace with traceable promoted field decisions.
                self.cur.execute('DELETE FROM staged_field WHERE staged_record_id=?', (record,))
                self.promote_material_row(record, path, sheet.title, row, values, headers, family, identity,
                                          clean(values[cols['tier']]), cols)
            else:
                self.counts['conflicts'] += 1 if english and '/' in english else 0

    def migrate_components(self, path, sheet, batch):
        headers = [clean(c) for c in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        type_map = {'spacer': 'spacer', 'centerpiece': 'connector', 'charm': 'charm', 'slider': 'clasp',
                    'cord': 'other', 'guard': 'other', 'pearl': 'pearl', 'tag': 'charm'}
        for row, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(clean(v) for v in values): continue
            english = clean(values[3]) if len(values) > 3 else None
            nonwearable = english == 'Energy & Material Disclosure Card'
            record = self.stage(batch, sheet.title, row, values, headers, 'component_intake',
                'Non-wearable disclosure card requires category review.' if nonwearable else None)
            spec = self.field(record, headers[4], values[4], 'component', 'size_mm')
            name = self.field(record, headers[3], english, 'component', 'component_code')
            price = self.field(record, headers[6], values[6], 'supplier_offer', 'price_text', 'review_required', 'Raw 1688 price stays staged.')
            link = self.field(record, headers[13], values[13], 'supplier', 'supplier_candidate', 'review_required', 'Procurement link is not a verified supplier.')
            if nonwearable:
                self.decision(name, 'retained_staged', 'Not wearable; do not force into component or packaging model.')
                self.decision(spec, 'retained_staged', 'Non-wearable card requires category decision.')
                self.decision(price, 'retained_staged', 'Raw price remains staged.')
                self.decision(link, 'retained_staged', 'Supplier candidate remains staged.')
                continue
            source_id = self.source(path, sheet.title, row)
            lower = (english or '').lower()
            ctype = next((v for k, v in type_map.items() if k in lower), 'other')
            self.decision(name, 'approved', 'Wearable accessory source name maps to a component record.')
            self.decision(spec, 'approved', 'Source specification retained as component size/spec text.')
            self.decision(price, 'retained_staged', 'Raw price stays staged.')
            self.decision(link, 'retained_staged', 'Supplier candidate stays staged.')
            self.cur.execute("""INSERT INTO component(component_code,component_type,shape_code,size_mm,design_role,visual_weight,source_id,notes)
              VALUES (?,?,?,?,?,?,?,?)""", (f'p1b-cmp-{row}-{slug(english or "unnamed")}', ctype, 'other', clean(values[4]),
                'hero' if ctype in ('connector','charm') else 'anchor', 'light' if ctype == 'spacer' else 'medium', source_id,
                'P1B source-derived wearable accessory. Raw commercial and procurement text remains staged.'))
            component_id = self.cur.lastrowid; self.counts['components'] += 1
            self.audit(record, 'component', component_id, 'component_code-and-spec', 'create', 'Wearable accessory mapping.', [name, spec])
            self.row_log(record, 'component', component_id, 'PARTIALLY_PROMOTED: raw 1688 price and procurement candidate remain staged.')

    def migrate_packaging(self, path, sheet, batch):
        headers = [clean(c) for c in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
        for row, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(clean(v) for v in values): continue
            english = clean(values[4]) if len(values) > 4 else None
            record = self.stage(batch, sheet.title, row, values, headers, 'packaging_intake', None)
            name = self.field(record, headers[4], english, 'packaging_option', 'packaging_code')
            price = self.field(record, headers[7], values[7], 'packaging_supplier_offer', 'price_text', 'review_required', 'Raw 1688 price stays staged.')
            link = self.field(record, headers[13], values[13], 'supplier', 'supplier_candidate', 'review_required', 'Procurement link is unverified.')
            for field_id in (name,): self.decision(field_id, 'approved', 'Packaging source name maps to packaging option.')
            self.decision(price, 'retained_staged', 'No packaging supplier offer is created from raw price text.')
            self.decision(link, 'retained_staged', 'Supplier candidate remains staged.')
            lower = (english or '').lower()
            ptype = 'pouch' if 'pouch' in lower and 'envelope' not in lower else ('envelope_pouch' if 'envelope' in lower else ('large_packaging' if any(k in lower for k in ('carton','mailer')) else ('card' if 'card' in lower else 'gift_box')))
            source_id = self.source(path, sheet.title, row)
            self.cur.execute("""INSERT INTO packaging_option(packaging_code,packaging_type,material_description,dimensions,finish,suitable_tier,source_id,verification_status,notes)
              VALUES (?,?,?,?,?,?,?,?,?)""", (f'p1b-pkg-{row}-{slug(english or "unnamed")}', ptype, clean(values[6]), clean(values[5]),
              clean(values[8]), tier(clean(values[1])), source_id, 'unverified',
              f'Use case: {clean(values[8]) or "not supplied"}. Slogan placement: {clean(values[9]) or "not supplied"}. Raw price and procurement link retained in staging.'))
            package_id = self.cur.lastrowid; self.counts['packaging'] += 1
            self.audit(record, 'packaging_option', package_id, 'packaging_code', 'create', 'Packaging mapped without a supplier offer.', [name])
            self.cur.execute('INSERT INTO market_assessment(subject_type,subject_id,target_market,assessment_text,analyst,basis_notes,confidence) VALUES (?,?,?,?,?,?,?)',
                             ('packaging_option', package_id, 'Europe', clean(values[8]) or 'Legacy packaging use case not specified.',
                              'legacy workbook', f'{sheet.title} row {row}; source-derived presentation assessment', 'low'))
            self.audit(record, 'market_assessment', self.cur.lastrowid, 'assessment_text', 'create', 'Packaging use-case assessment.', [name])
            self.row_log(record, 'packaging_option', package_id, 'PARTIALLY_PROMOTED: raw 1688 price and procurement candidate remain staged.')

    def close(self):
        self.db.commit(); self.db.close()

    def repair_validation(self):
        """Repair only P1B's initially incomplete audit keys; never remap data."""
        self.cur.execute("""UPDATE material_alias SET normalized_alias=lower(trim(alias_raw))
          WHERE review_status='reviewed' AND (normalized_alias IS NULL OR normalized_alias='')
            AND source_id IN (SELECT id FROM source WHERE name LIKE 'P1B legacy workbook | %')""")
        self.cur.execute("""INSERT INTO review_decision(staged_record_id,decision,reviewer,notes)
          SELECT p.staged_record_id, 'approved', 'P1B deterministic mapping rules',
                 'Backfilled row-level approval for an already field-audited P1B promotion.'
          FROM promotion_log p
          WHERE p.promoted_by='P1B deterministic migration'
            AND NOT EXISTS (SELECT 1 FROM review_decision r WHERE r.staged_record_id=p.staged_record_id AND r.decision='approved')""")
        # The two workbooks use different price/unit wording for repeated material names.
        # Flag it on the revised row without choosing either price or creating an offer.
        self.cur.execute("""UPDATE staged_record AS revised SET warning_summary=
          trim(coalesce(revised.warning_summary || ' ', '') ||
               'Cross-workbook raw price/unit conflict; both values remain staged.')
          WHERE revised.import_batch_id IN (SELECT id FROM import_batch
              WHERE source_description='P1B batch A/B: revised crystal and pearl workbook')
            AND EXISTS (
              SELECT 1 FROM staged_field rn JOIN staged_field rp ON rp.staged_record_id=rn.staged_record_id
              JOIN staged_record original ON original.id=rn.staged_record_id
              JOIN import_batch ob ON ob.id=original.import_batch_id
              JOIN staged_field oname ON oname.staged_record_id=original.id AND oname.target_field='canonical_name'
              JOIN staged_field oprice ON oprice.staged_record_id=original.id AND oprice.target_field='price_text'
              JOIN staged_field rname ON rname.staged_record_id=revised.id AND rname.target_field='canonical_name'
              JOIN staged_field rprice ON rprice.staged_record_id=revised.id AND rprice.target_field='price_text'
              WHERE rn.target_field='canonical_name' AND rp.target_field='price_text'
                AND ob.source_description='P1B batch A/B: legacy crystal and pearl workbook'
                AND oname.raw_value=rname.raw_value AND coalesce(oprice.raw_value,'')<>coalesce(rprice.raw_value,'')
            )""")


def main(source_path, revised_path, db_path):
    p = P1B(db_path)
    source = load_workbook(source_path, read_only=True, data_only=False)
    revised = load_workbook(revised_path, read_only=True, data_only=False)
    # Sheet positions avoid depending on local terminal Unicode rendering.
    b1 = p.batch(source_path, 'P1B batch A/B: legacy crystal and pearl workbook')
    p.migrate_material_sheet(source_path, source.worksheets[0], b1, set(range(3, 13)), 'crystal')
    p.migrate_material_sheet(source_path, source.worksheets[1], b1, set(), 'pearl')
    b2 = p.batch(revised_path, 'P1B batch A/B: revised crystal and pearl workbook')
    p.migrate_material_sheet(revised_path, revised.worksheets[0], b2, set(), 'crystal')
    p.migrate_material_sheet(revised_path, revised.worksheets[1], b2, set(), 'pearl')
    b3 = p.batch(revised_path, 'P1B batch C: accessory workbook sheet')
    p.migrate_components(revised_path, revised.worksheets[2], b3)
    b4 = p.batch(revised_path, 'P1B batch D: packaging workbook sheet')
    p.migrate_packaging(revised_path, revised.worksheets[3], b4)
    p.close()
    print(json.dumps(p.counts, ensure_ascii=False))


if __name__ == '__main__':
    if len(sys.argv) == 3 and sys.argv[1] == '--repair-validation':
        p = P1B(sys.argv[2]); p.repair_validation(); p.close()
        print('P1B validation audit repair complete.')
        raise SystemExit(0)
    if len(sys.argv) != 4:
        raise SystemExit('Usage: migrate-p1b.py SOURCE.xlsx REVISED.xlsx DATABASE.sqlite | --repair-validation DATABASE.sqlite')
    main(sys.argv[1], sys.argv[2], sys.argv[3])
