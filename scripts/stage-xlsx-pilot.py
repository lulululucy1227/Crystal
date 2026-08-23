"""Read-only P1A-R pilot staging for the inspected source workbook.

This utility stages at most ten rows. It never writes canonical material,
variant, component, supplier, offer, narrative, or assessment records.
"""
import json
import sqlite3
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

MAX_ROWS = 10

def clean(value):
    if value is None:
        return None
    return str(value).strip() or None

def add_field(cursor, record_id, column, raw, target_entity, target_field, status='review_required', warning=None):
    raw = clean(raw)
    cursor.execute(
        """INSERT INTO staged_field(staged_record_id,source_column,raw_value,normalized_value,target_entity,target_field,field_status,warning_or_error)
           VALUES (?,?,?,?,?,?,?,?)""",
        (record_id, column, raw, raw, target_entity, target_field, status, warning),
    )

def main(workbook_path, db_path):
    workbook_path = Path(workbook_path)
    wb = load_workbook(workbook_path, read_only=True, data_only=False)
    sheet = wb['水晶'] if '水晶' in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    next(rows, None)  # merged title
    headers = list(next(rows, []))
    # The workbook has a merged title row, then a stable P1A audit layout:
    # C Chinese material name, D English name, E quality, G tier, H EU note,
    # I narrative, J size, K price text, L supplier candidate. Positional
    # mapping avoids making data safety depend on terminal Unicode rendering.
    columns = {'chinese_name': 2, 'english_name': 3, 'quality': 4, 'tier': 6,
               'market': 7, 'narrative': 8, 'size': 9, 'price': 10, 'supplier': 11}
    if len(headers) <= max(columns.values()):
        raise SystemExit('Workbook does not have the audited P1A material-column layout.')

    db = sqlite3.connect(db_path)
    cur = db.cursor()
    cur.execute(
        """INSERT INTO import_batch(source_file,source_format,source_description,imported_by,batch_status)
           VALUES (?,?,?,?,?)""",
        (str(workbook_path), 'xlsx', 'P1A-R read-only maximum-ten-row material pilot', 'P1A-R pilot reader', 'ready_for_review'),
    )
    batch_id = cur.lastrowid
    review_rows = []
    for excel_row, values in enumerate(rows, start=3):
        chinese_name = clean(values[columns['chinese_name']]) if len(values) > columns['chinese_name'] else None
        english_name = clean(values[columns['english_name']]) if len(values) > columns['english_name'] else None
        if not chinese_name and not english_name:
            continue
        raw = {headers[i] or f'column_{i + 1}': clean(value) for i, value in enumerate(values) if clean(value) is not None}
        warnings = [
            'Real source row is staged only; explicit human approval is required before any canonical write.',
            'Material identity, alias equivalence, grade, price unit, and supplier identity require review.',
        ]
        cur.execute(
            """INSERT INTO staged_record(import_batch_id,source_sheet,source_row,raw_record_json,target_entity,validation_status,warning_summary)
               VALUES (?,?,?,?,?,?,?)""",
            (batch_id, sheet.title, excel_row, json.dumps(raw, ensure_ascii=False), 'material_intake', 'review_required', ' '.join(warnings)),
        )
        record_id = cur.lastrowid
        add_field(cur, record_id, headers[columns['chinese_name']], chinese_name, 'material_alias', 'alias_raw', warning='Chinese source name is only a proposed identity alias.')
        add_field(cur, record_id, headers[columns['english_name']], english_name, 'material', 'canonical_name', warning='English source name is a proposal, not an automatic canonical material.')
        add_field(cur, record_id, headers[columns['quality']], values[columns['quality']], 'material_variant', 'quality_description', warning='Free text must be split into separately reviewed variant attributes.')
        add_field(cur, record_id, headers[columns['tier']], values[columns['tier']], 'material_variant', 'source_tier_label', warning='Low/mid/high is a commercial positioning label, never a material identity tier.')
        add_field(cur, record_id, headers[columns['size']], values[columns['size']], 'material_variant', 'size_range_mm', warning='Size range does not identify a unique component.')
        add_field(cur, record_id, headers[columns['price']], values[columns['price']], 'supplier_offer', 'price_text', warning='Price is a text range; currency, unit and supplier quote need review.')
        add_field(cur, record_id, headers[columns['supplier']], values[columns['supplier']], 'supplier', 'supplier_candidate', warning='Supplier string is a candidate only and has not been verified.')
        add_field(cur, record_id, headers[columns['narrative']], values[columns['narrative']], 'material_narrative', 'statement', warning='Energy/symbolism is a narrative, not a mineralogical fact or medical claim.')
        add_field(cur, record_id, headers[columns['market']], values[columns['market']], 'market_assessment', 'assessment_text', warning='Workbook market suitability is an internal assessment, not market evidence.')
        review_rows.append({
            'staged_record_id': record_id, 'source_file': str(workbook_path), 'sheet': sheet.title, 'row': excel_row,
            'raw_material_name': chinese_name, 'proposed_canonical_material': english_name,
            'proposed_aliases': [x for x in (chinese_name, english_name) if x],
            'proposed_variant': clean(values[columns['quality']]), 'component_information': clean(values[columns['size']]),
            'source_tier': clean(values[columns['tier']]), 'price_parse': clean(values[columns['price']]),
            'supplier_candidate': clean(values[columns['supplier']]),
            'market_assessment_candidate': clean(values[columns['market']]),
            'narrative_candidate': clean(values[columns['narrative']]),
            'confidence': 'low', 'review_required_reasons': warnings,
        })
        if len(review_rows) == MAX_ROWS:
            break
    db.commit()
    db.close()
    print(json.dumps({'batch_id': batch_id, 'staged_rows': review_rows, 'generated_at': datetime.now(timezone.utc).isoformat()}, ensure_ascii=False, indent=2))

if __name__ == '__main__':
    if len(sys.argv) != 3:
        raise SystemExit('Usage: stage-xlsx-pilot.py SOURCE.xlsx DATABASE.sqlite')
    main(sys.argv[1], sys.argv[2])
